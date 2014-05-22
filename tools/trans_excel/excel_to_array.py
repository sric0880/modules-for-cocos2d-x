from openpyxl import load_workbook
import plistlib
import sys
import json
from sqlite3 import *

def work(excel_file_name):
	wb = load_workbook(excel_file_name)
	sheetnames = wb.get_sheet_names()
	ret=[]
	for name in sheetnames:
		ws = wb.get_sheet_by_name(name)
		arr=[]
		header = ws.rows[0]
		for rid in range(1,ws.get_highest_row()):
			row = ws.rows[rid]
			row_dic={}
			i = 0
			for cell in row:
				row_dic[header[i].value]=cell.value
				i=i+1
			arr.append(row_dic)
		ret.append((ws.title,arr,header))
	return ret

def fill_work(raw_array): #fill the blank cell, use in sqlite
	arr_len = len(raw_array)
	i = 0
	while i < arr_len:
		for key, value in raw_array[i].items():
			if value == '':
				raw_array[i][key] = raw_array[i-1][key]
		i = i + 1
	return raw_array

def merge_work(raw_array, col_index, header): #merge the blank cell
	good_array=[]
	arr_len = len(raw_array)
	keyword = header[col_index].value
	i = 0
	while i < arr_len:
		arr=[]
		not_blank_cell_var = raw_array[i][keyword]
		arr.append({ key : value for key, value in raw_array[i].items() if key != keyword})
		while 1:
			i = i + 1
			if(i == arr_len):
				break
			blank_cell_var = raw_array[i][keyword]
			if(blank_cell_var != ''):
				break
			arr.append({ key : value for key, value in raw_array[i].items() if key != keyword})
		if(len(arr) == 1):
			good_array.append(raw_array[i - 1])
		else:
			good_array.append(merge_work(arr,col_index+1,header))
	return good_array


if __name__=='__main__':
	if(len(sys.argv)<3):
		print('Usage: python excel_to_array.py [excel_file_name] [dest_file_type:plist|json|sqlite]')
	else:
		data = work(sys.argv[1])
		if(sys.argv[2].lower()=='json'):
			for (file_name, raw_array, header) in data:
				good_arr = merge_work(raw_array,0, header)
				f = open(file_name+'.json','w')
				c = json.dumps(good_arr, ensure_ascii=False, indent=2)
				if(isinstance(c, unicode)):
					 f.write(c.encode('utf8'))
				else:
					f.write(c)
				f.close()
		elif(sys.argv[2].lower()=='plist'):
			for (file_name, raw_array, header) in data:
				good_arr = merge_work(raw_array,0, header)
				plistlib.writePlist(good_arr, file_name+'.plist')
		elif(sys.argv[2].lower()=='sqlite'):
			conn = connect('rename_it.sqlite')
			with conn:
				cur = conn.cursor()
				for (file_name, raw_array, header) in data:
					s = ','.join([cell.value for cell in header])
					sql = 'CREATE TABLE %s(%s)'%(file_name,s)
					print(sql)
					cur.execute(sql)
					good_arr = fill_work(raw_array)
					for item in good_arr:
						new_v=[]
						for v in [item[cell.value] for cell in header]:
							if(isinstance(v, unicode)):
								new_v.append('\''+v.encode('utf8')+'\'')
							elif(isinstance(v,str)):
								new_v.append('\''+v+'\'')
							else:
								new_v.append(str(v))
						s = ','.join(new_v)
						sql = 'INSERT INTO %s VALUES(%s)'%(file_name, s) #'index' is key word in sqlite so don't use it
						print(sql)
						cur.execute(sql)
					
		else:
			print("Error: wrong dest_file_type")